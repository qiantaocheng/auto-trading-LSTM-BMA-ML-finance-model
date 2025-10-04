#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""快速检查样本数量问题"""

import openpyxl

# 读取最新的分析结果
wb = openpyxl.load_workbook('result/bma_analysis_20251003_153219.xlsx', data_only=True)

# 检查模型信息sheet
if '模型信息' in wb.sheetnames:
    ws = wb['模型信息']
    print("=== 模型信息 ===")
    for row in ws.iter_rows(min_row=1, max_row=40, values_only=True):
        if row[0]:  # 只打印非空行
            print(f"{row[0]}: {row[1] if len(row) > 1 else ''}")
    print()

# 检查预测结果sheet
if '预测结果' in wb.sheetnames:
    ws = wb['预测结果']
    print(f"\n=== 预测结果行数 ===")
    print(f"总行数 (含表头): {ws.max_row}")
    print(f"有效预测数: {ws.max_row - 1}")

wb.close()
