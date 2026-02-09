#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Direct Prediction Excel Ranking Report (EMA-smoothed)
=====================================================
Generates Excel ranking report using EMA-smoothed predictions (default: 4-day window, \u03b2=0.33).

Contents:
- Smoothed scores for the last few sessions (with raw backups)
- Rankings and score deltas after smoothing
- Model breakdowns (LambdaRank, CatBoost, ElasticNet, XGBoost)
"""

import sys
import argparse
import json
import os
from pathlib import Path
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import logging

sys.path.insert(0, str(Path(__file__).parent.parent))

from bma_models.model_registry import load_manifest
from bma_models.量化模型_bma_ultra_enhanced import UltraEnhancedQuantitativeModel
from bma_models.simple_25_factor_engine import T10_ALPHA_FACTORS as DEFAULT_FEATURES

EXCLUDED_FEATURES = {"obv_divergence"}

def _filter_disallowed_features(features, logger, context):
    filtered = [f for f in features if f not in EXCLUDED_FEATURES]
    removed = sorted(set(features) & EXCLUDED_FEATURES)
    if removed:
        logger.warning("Removing disabled features %s from %s", removed, context)
    if not filtered:
        raise ValueError(f"{context} is empty after filtering disallowed features")
    return filtered


logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

feature_list = _filter_disallowed_features(list(DEFAULT_FEATURES), logger, "direct prediction default features")
os.environ["BMA_FEATURE_OVERRIDES"] = json.dumps({
    "elastic_net": feature_list,
    "xgboost": feature_list,
    "catboost": feature_list,
    "lightgbm_ranker": feature_list,
    "lambdarank": feature_list,
})
logger.info(f"Using default feature set ({len(feature_list)} factors) with obv_divergence disabled: {feature_list}")



def calculate_ewma_smoothed_scores(predictions_df: pd.DataFrame,
                                   weights: tuple = (0.41, 0.28, 0.19, 0.12),
                                   use_half_life: bool = False,
                                   half_life_days: float = 3.0) -> pd.DataFrame:
    """Calculate EMA-smoothed scores across an arbitrary lookback window."""
    if not isinstance(predictions_df.index, pd.MultiIndex):
        raise ValueError("predictions_df must have MultiIndex (date, ticker)")
    if 'score' not in predictions_df.columns:
        raise ValueError("predictions_df must have 'score' column")

    dates = sorted(predictions_df.index.get_level_values('date').unique())
    if not dates:
        return predictions_df.copy()

    base_weights = tuple(float(w) for w in weights if float(w) > 0)
    if not base_weights:
        raise ValueError('At least one positive weight is required for EMA smoothing')

    if use_half_life:
        window_size = max(len(base_weights), 2)
        alpha = 1 - np.exp(-np.log(2) / max(half_life_days, 1.0))
        decay = [alpha * ((1 - alpha) ** i) for i in range(window_size)]
        total = sum(decay)
        weights = tuple(w / total for w in decay)
        logger.info(f"Using half-life={half_life_days} days, calculated weights: {weights}")
    else:
        total = sum(base_weights)
        weights = tuple(w / total for w in base_weights)
        logger.info(f"Using fixed weights: {weights}")

    window = len(weights)
    if len(dates) < window:
        logger.warning(f"Only {len(dates)} day(s) available; EMA window={window}. Will renormalize with available data")

    result_df = predictions_df.copy()
    result_df['score_raw'] = result_df['score'].copy()
    result_df['score_smoothed'] = np.nan

    for ticker in predictions_df.index.get_level_values('ticker').unique():
        for idx, current_date in enumerate(dates):
            scores = []
            for offset in range(window):
                pos = idx - offset
                if pos < 0:
                    break
                candidate_date = dates[pos]
                key = (candidate_date, ticker)
                if key in predictions_df.index:
                    scores.append(predictions_df.loc[key, 'score'])
            if not scores:
                smoothed = predictions_df.loc[(current_date, ticker), 'score'] if (current_date, ticker) in predictions_df.index else np.nan
            else:
                used_weights = np.array(weights[:len(scores)], dtype=np.float64)
                used_weights /= used_weights.sum()
                smoothed = float(np.dot(used_weights, scores))
            if (current_date, ticker) in result_df.index:
                result_df.loc[(current_date, ticker), 'score_smoothed'] = smoothed

    result_df['score'] = result_df['score_smoothed'].fillna(result_df['score_raw'])
    return result_df


def generate_excel_ranking_report(predictions_df: pd.DataFrame,
                                 output_path: str,
                                 model_name: str = "MetaRankerStacker",
                                 top_n: int = 20) -> None:
    """
    Generate Excel ranking report with Top N predictions (LambdaRank, CatBoost, MetaRankerStacker)
    
    Args:
        predictions_df: DataFrame with MultiIndex (date, ticker) and score columns
        output_path: Path to output Excel file
        model_name: Name of the model for report title
        top_n: Number of top predictions to show (default: 20)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    logger.info(f"📊 Generating Excel ranking report: {output_path}")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking Report"
    
    # Get latest date
    dates = sorted(predictions_df.index.get_level_values('date').unique())
    latest_date = dates[-1] if len(dates) > 0 else None
    
    # Title
    title_cell = ws['A1']
    title_cell.value = f"Top {top_n} Predictions: All Models (Raw Scores, No EMA)"
    title_cell.font = Font(size=16, bold=True)
    ws.merge_cells('A1:J1')

    # Date info
    date_cell = ws['A2']
    date_cell.value = f"Report Date: {latest_date.strftime('%Y-%m-%d') if latest_date else 'N/A'}"
    date_cell.font = Font(size=12)
    ws.merge_cells('A2:J2')

    note_cell = ws['A3']
    note_cell.value = f"Note: Top {top_n} list uses EMA-smoothed scores (default window 4, beta=0.33)."
    note_cell.font = Font(size=10, italic=True)
    ws.merge_cells('A3:J3')

    # Headers
    headers = ['Rank', 'Ticker', f'{model_name} Score', 'LambdaRank Score', 'CatBoost Score', 'ElasticNet Score', 'XGBoost Score', 'Score (Yesterday)', 'Score Change']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Get latest date predictions
    if latest_date:
        try:
            latest_predictions = predictions_df.xs(latest_date, level='date', drop_level=False)
        except KeyError:
            logger.warning(f"Date {latest_date} not found in predictions, using all available data")
            latest_predictions = predictions_df.groupby(level='ticker').last()
        
        # Ensure we have a smoothed score column
        if 'score' not in latest_predictions.columns:
            if 'score_smoothed' in latest_predictions.columns:
                latest_predictions['score'] = latest_predictions['score_smoothed']
            elif 'score_raw' in latest_predictions.columns:
                latest_predictions['score'] = latest_predictions['score_raw']
            else:
                logger.warning("No score column found, using first numeric column")
                numeric_cols = latest_predictions.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    latest_predictions['score'] = latest_predictions[numeric_cols[0]]
                else:
                    raise ValueError("No score column found in predictions")

        latest_predictions = latest_predictions.sort_values('score', ascending=False)
        
        # 🔥 Limit to Top N only
        latest_predictions = latest_predictions.head(top_n)
        logger.info(f"📊 Showing Top {top_n} predictions only (out of {len(latest_predictions)} total)")
        
        # Get previous days for comparison
        prev_dates = dates[-2:] if len(dates) >= 2 else dates  # Only need yesterday for comparison
        
        row = 5
        for rank, (idx, row_data) in enumerate(latest_predictions.iterrows(), start=1):
            ticker = idx[1] if isinstance(idx, tuple) else idx
            
            # Use EMA-smoothed meta ranker scores
            score_meta_stacker = row_data.get('score', np.nan)
            if np.isnan(score_meta_stacker):
                score_meta_stacker = row_data.get('score_raw', np.nan)

            score_lambdarank = row_data.get('score_lambdarank', np.nan)
            score_catboost = row_data.get('score_catboost', np.nan)
            score_elastic = row_data.get('score_elastic', np.nan)
            score_xgb = row_data.get('score_xgb', np.nan)

            scores_yesterday = np.nan
            if len(prev_dates) >= 2:
                yesterday_date = prev_dates[-2]
                try:
                    if (yesterday_date, ticker) in predictions_df.index:
                        scores_yesterday = predictions_df.loc[(yesterday_date, ticker), 'score']
                except (KeyError, IndexError):
                    pass
            
            # Calculate score change
            score_change = np.nan
            if not np.isnan(score_meta_stacker) and not np.isnan(scores_yesterday):
                score_change = score_meta_stacker - scores_yesterday
            
            # Write row
            ws.cell(row=row, column=1).value = rank
            ws.cell(row=row, column=2).value = ticker
            ws.cell(row=row, column=3).value = score_meta_stacker if not np.isnan(score_meta_stacker) else None
            ws.cell(row=row, column=4).value = score_lambdarank if not np.isnan(score_lambdarank) else None
            ws.cell(row=row, column=5).value = score_catboost if not np.isnan(score_catboost) else None
            ws.cell(row=row, column=6).value = score_elastic if not np.isnan(score_elastic) else None
            ws.cell(row=row, column=7).value = score_xgb if not np.isnan(score_xgb) else None
            ws.cell(row=row, column=8).value = scores_yesterday if not np.isnan(scores_yesterday) else None
            ws.cell(row=row, column=9).value = score_change if not np.isnan(score_change) else None
            
            # Format cells
            for col in range(1, 10):  # 9 columns: Rank, Ticker, MetaStacker, LambdaRank, CatBoost, ElasticNet, XGBoost, Yesterday, Change
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center" if col == 1 else "left" if col == 2 else "right", vertical="center")
                if col == 1:  # Rank
                    cell.font = Font(bold=True)
                if col == 9 and not np.isnan(score_change):  # Score change (column 9)
                    if score_change > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif score_change < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
    else:
        logger.warning("No dates found in predictions")
    
    # Auto-adjust column widths
    for col in range(1, 10):  # 9 columns
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add CatBoost Top 20 sheet
    if latest_date and 'score_catboost' in latest_predictions.columns:
        ws_catboost = wb.create_sheet("CatBoost Top20")
        
        # Title
        title_cell = ws_catboost['A1']
        title_cell.value = f"CatBoost Top {top_n} Predictions"
        title_cell.font = Font(size=16, bold=True)
        ws_catboost.merge_cells('A1:C1')
        
        # Date info
        date_cell = ws_catboost['A2']
        date_cell.value = f"Report Date: {latest_date.strftime('%Y-%m-%d') if latest_date else 'N/A'}"
        date_cell.font = Font(size=12)
        ws_catboost.merge_cells('A2:C2')
        
        # Headers
        headers = ['Rank', 'Ticker', 'CatBoost Score']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_catboost.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get CatBoost top 20
        catboost_top20 = latest_predictions.nlargest(top_n, 'score_catboost')[['score_catboost']].copy()
        catboost_top20 = catboost_top20.sort_values('score_catboost', ascending=False)
        
        row = 4
        for rank, (idx, row_data) in enumerate(catboost_top20.iterrows(), start=1):
            ticker = idx[1] if isinstance(idx, tuple) else idx
            score = row_data['score_catboost']
            
            ws_catboost.cell(row=row, column=1).value = rank
            ws_catboost.cell(row=row, column=2).value = ticker
            ws_catboost.cell(row=row, column=3).value = score if not np.isnan(score) else None
            
            # Format cells
            for col in range(1, 4):
                cell = ws_catboost.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center" if col == 1 else "left" if col == 2 else "right", vertical="center")
                if col == 1:  # Rank
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 4):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_catboost[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_catboost.column_dimensions[column].width = adjusted_width
    
    # Add LambdaRanker Top 20 sheet
    if latest_date and 'score_lambdarank' in latest_predictions.columns:
        ws_lambdarank = wb.create_sheet("LambdaRanker Top20")
        
        # Title
        title_cell = ws_lambdarank['A1']
        title_cell.value = f"LambdaRanker Top {top_n} Predictions"
        title_cell.font = Font(size=16, bold=True)
        ws_lambdarank.merge_cells('A1:C1')
        
        # Date info
        date_cell = ws_lambdarank['A2']
        date_cell.value = f"Report Date: {latest_date.strftime('%Y-%m-%d') if latest_date else 'N/A'}"
        date_cell.font = Font(size=12)
        ws_lambdarank.merge_cells('A2:C2')
        
        # Headers
        headers = ['Rank', 'Ticker', 'LambdaRanker Score']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_lambdarank.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get LambdaRanker top 20
        lambdarank_top20 = latest_predictions.nlargest(top_n, 'score_lambdarank')[['score_lambdarank']].copy()
        lambdarank_top20 = lambdarank_top20.sort_values('score_lambdarank', ascending=False)
        
        row = 4
        for rank, (idx, row_data) in enumerate(lambdarank_top20.iterrows(), start=1):
            ticker = idx[1] if isinstance(idx, tuple) else idx
            score = row_data['score_lambdarank']
            
            ws_lambdarank.cell(row=row, column=1).value = rank
            ws_lambdarank.cell(row=row, column=2).value = ticker
            ws_lambdarank.cell(row=row, column=3).value = score if not np.isnan(score) else None
            
            # Format cells
            for col in range(1, 4):
                cell = ws_lambdarank.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center" if col == 1 else "left" if col == 2 else "right", vertical="center")
                if col == 1:  # Rank
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 4):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_lambdarank[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_lambdarank.column_dimensions[column].width = adjusted_width
    
    # Add ElasticNet Top 20 sheet
    if latest_date and 'score_elastic' in latest_predictions.columns:
        ws_elastic = wb.create_sheet("ElasticNet Top20")
        
        # Title
        title_cell = ws_elastic['A1']
        title_cell.value = f"ElasticNet Top {top_n} Predictions"
        title_cell.font = Font(size=16, bold=True)
        ws_elastic.merge_cells('A1:C1')
        
        # Date info
        date_cell = ws_elastic['A2']
        date_cell.value = f"Report Date: {latest_date.strftime('%Y-%m-%d') if latest_date else 'N/A'}"
        date_cell.font = Font(size=12)
        ws_elastic.merge_cells('A2:C2')
        
        # Headers
        headers = ['Rank', 'Ticker', 'ElasticNet Score']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_elastic.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get ElasticNet top 20
        elastic_top20 = latest_predictions.nlargest(top_n, 'score_elastic')[['score_elastic']].copy()
        elastic_top20 = elastic_top20.sort_values('score_elastic', ascending=False)
        
        row = 4
        for rank, (idx, row_data) in enumerate(elastic_top20.iterrows(), start=1):
            ticker = idx[1] if isinstance(idx, tuple) else idx
            score = row_data['score_elastic']
            
            ws_elastic.cell(row=row, column=1).value = rank
            ws_elastic.cell(row=row, column=2).value = ticker
            ws_elastic.cell(row=row, column=3).value = score if not np.isnan(score) else None
            
            # Format cells
            for col in range(1, 4):
                cell = ws_elastic.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center" if col == 1 else "left" if col == 2 else "right", vertical="center")
                if col == 1:  # Rank
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 4):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_elastic[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_elastic.column_dimensions[column].width = adjusted_width
    
    # Add XGBoost Top 20 sheet
    if latest_date and 'score_xgb' in latest_predictions.columns:
        ws_xgb = wb.create_sheet("XGBoost Top20")
        
        # Title
        title_cell = ws_xgb['A1']
        title_cell.value = f"XGBoost Top {top_n} Predictions"
        title_cell.font = Font(size=16, bold=True)
        ws_xgb.merge_cells('A1:C1')
        
        # Date info
        date_cell = ws_xgb['A2']
        date_cell.value = f"Report Date: {latest_date.strftime('%Y-%m-%d') if latest_date else 'N/A'}"
        date_cell.font = Font(size=12)
        ws_xgb.merge_cells('A2:C2')
        
        # Headers
        headers = ['Rank', 'Ticker', 'XGBoost Score']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_xgb.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get XGBoost top 20
        xgb_top20 = latest_predictions.nlargest(top_n, 'score_xgb')[['score_xgb']].copy()
        xgb_top20 = xgb_top20.sort_values('score_xgb', ascending=False)
        
        row = 4
        for rank, (idx, row_data) in enumerate(xgb_top20.iterrows(), start=1):
            ticker = idx[1] if isinstance(idx, tuple) else idx
            score = row_data['score_xgb']
            
            ws_xgb.cell(row=row, column=1).value = rank
            ws_xgb.cell(row=row, column=2).value = ticker
            ws_xgb.cell(row=row, column=3).value = score if not np.isnan(score) else None
            
            # Format cells
            for col in range(1, 4):
                cell = ws_xgb.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center" if col == 1 else "left" if col == 2 else "right", vertical="center")
                if col == 1:  # Rank
                    cell.font = Font(bold=True)
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 4):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws_xgb[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_xgb.column_dimensions[column].width = adjusted_width
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Summary Statistics"
    ws_summary['A1'].font = Font(size=14, bold=True)
    
    if latest_date and len(latest_predictions) > 0:
        ws_summary['A3'] = f"Total Tickers (Top {top_n} shown)"
        ws_summary['B3'] = top_n
        ws_summary['A4'] = f"Top {top_n} Avg Score ({model_name})"
        ws_summary['B4'] = latest_predictions.head(top_n)['score'].mean() if len(latest_predictions) >= top_n else latest_predictions['score'].mean()
        if 'score_lambdarank' in latest_predictions.columns:
            ws_summary['A5'] = f"Top {top_n} Avg Score (LambdaRank)"
            ws_summary['B5'] = latest_predictions.head(top_n)['score_lambdarank'].mean() if len(latest_predictions) >= top_n else latest_predictions['score_lambdarank'].mean()
        if 'score_catboost' in latest_predictions.columns:
            ws_summary['A6'] = f"Top {top_n} Avg Score (CatBoost)"
            ws_summary['B6'] = latest_predictions.head(top_n)['score_catboost'].mean() if len(latest_predictions) >= top_n else latest_predictions['score_catboost'].mean()
        if 'score_elastic' in latest_predictions.columns:
            ws_summary['A7'] = f"Top {top_n} Avg Score (ElasticNet)"
            ws_summary['B7'] = latest_predictions.head(top_n)['score_elastic'].mean() if len(latest_predictions) >= top_n else latest_predictions['score_elastic'].mean()
        if 'score_xgb' in latest_predictions.columns:
            ws_summary['A8'] = f"Top {top_n} Avg Score (XGBoost)"
            ws_summary['B8'] = latest_predictions.head(top_n)['score_xgb'].mean() if len(latest_predictions) >= top_n else latest_predictions['score_xgb'].mean()
    
    # Save workbook
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info(f"✅ Excel report saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Direct prediction with EWMA smoothing and Excel ranking report")
    parser.add_argument("--tickers", type=str, required=True, help="Comma-separated ticker symbols")
    parser.add_argument("--snapshot", type=str, default="latest", help="Snapshot ID or 'latest'")
    parser.add_argument("--days", type=int, default=4, help="Number of days to fetch for EWMA (default: 4)")
    parser.add_argument("--weights", type=str, default="0.41,0.28,0.19,0.12", help="EMA weights (most recent first) or 'half-life'")
    parser.add_argument("--half-life", type=float, default=None, help="Half-life in days (overrides weights if provided)")
    parser.add_argument("--output", type=str, default="results/ewma_ranking_report.xlsx", help="Output Excel file path")
    parser.add_argument("--as-of-date", type=str, default=None, help="Prediction date (YYYY-MM-DD), default: today")
    
    args = parser.parse_args()
    
    # Parse tickers
    tickers = [t.strip().upper() for t in args.tickers.split(',') if t.strip()]
    if not tickers:
        raise ValueError("No tickers provided")
    
    logger.info("=" * 100)
    logger.info("DIRECT PREDICTION WITH EWMA SMOOTHING")
    logger.info("=" * 100)
    logger.info(f"Tickers: {', '.join(tickers)}")
    logger.info(f"Days: {args.days}")
    
    # Parse EMA weights / half-life settings
    use_half_life = False
    half_life_days = args.half_life if args.half_life else None
    weights_tuple = (0.41, 0.28, 0.19, 0.12)

    weights_arg = args.weights.lower().strip() if isinstance(args.weights, str) else ''
    if args.half_life:
        use_half_life = True
        logger.info(f"Using half-life: {half_life_days} days")
    elif weights_arg in {'half-life', 'halflife'}:
        use_half_life = True
        half_life_days = 3.0
        logger.info("Using half-life: 3.0 days (default)")
    else:
        try:
            weight_list = [float(w.strip()) for w in args.weights.split(',') if w.strip()]
            if len(weight_list) >= 2:
                weights_tuple = tuple(weight_list)
                logger.info(f"Using fixed weights: {weights_tuple}")
            else:
                logger.warning("Invalid weights format, using default 4-day profile")
        except Exception as e:
            logger.warning(f"Failed to parse weights: {e}, using default 4-day profile")

    if use_half_life:
        ema_window = max(int(round(args.days)), 2)
        weights_tuple = tuple([1.0] * ema_window)
    else:
        ema_window = len(weights_tuple)

    if args.days < ema_window:
        logger.info(f"Expanding --days from {args.days} to {ema_window} to match EMA window")
        args.days = ema_window
    
    # Determine date range
    if args.as_of_date:
        end_date = pd.to_datetime(args.as_of_date)
    else:
        end_date = pd.Timestamp.today()
    
    # Need at least 3 days of data for EWMA (plus lookback for features)
    # Fetch data from (end_date - days - lookback) to end_date
    MIN_LOOKBACK_DAYS = 280  # For feature calculation
    start_date = (end_date - pd.Timedelta(days=args.days + MIN_LOOKBACK_DAYS)).strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')
    
    logger.info(f"Date range: {start_date} to {end_date_str}")
    logger.info(f"Prediction period: Last {args.days} days")
    
    # Initialize model
    model = UltraEnhancedQuantitativeModel()
    
    # Get predictions for last N days using predict_with_snapshot
    logger.info("📡 Fetching data and generating predictions...")
    
    all_predictions = []
    
    # Get predictions for all days at once (more efficient)
    # Fetch data for the entire period, then extract predictions for each day
    logger.info(f"🔮 Fetching data and generating predictions for last {args.days} days...")
    
    try:
        # Use predict_with_snapshot with auto-fetch for the entire period
        results = model.predict_with_snapshot(
            feature_data=None,  # Auto-fetch
            snapshot_id=None if args.snapshot == "latest" else args.snapshot,
            universe_tickers=tickers,
            as_of_date=end_date,
            prediction_days=args.days  # Get predictions for all days
        )
        
        if results.get('success', False):
            predictions = results.get('predictions')
            if predictions is not None and len(predictions) > 0:
                # Convert to DataFrame
                if isinstance(predictions, pd.Series):
                    pred_df = predictions.to_frame('score')
                elif isinstance(predictions, pd.DataFrame):
                    pred_df = predictions.copy()
                    if 'score' not in pred_df.columns and len(pred_df.columns) == 1:
                        pred_df.columns = ['score']
                else:
                    logger.error(f"Unexpected predictions type: {type(predictions)}")
                    raise ValueError("Predictions must be Series or DataFrame")
                
                # Ensure MultiIndex with date and ticker
                if not isinstance(pred_df.index, pd.MultiIndex):
                    logger.warning("Predictions don't have MultiIndex, attempting to infer dates...")
                    # If no MultiIndex, assume all predictions are for the latest date
                    pred_df.index = pd.MultiIndex.from_arrays(
                        [[end_date] * len(pred_df), pred_df.index],
                        names=['date', 'ticker']
                    )
                
                # Check if we have multiple dates
                dates_in_pred = sorted(pred_df.index.get_level_values('date').unique())
                logger.info(f"✅ Found predictions for {len(dates_in_pred)} dates: {dates_in_pred}")
                
                # If we only have one date but need multiple days, try to get historical predictions
                if len(dates_in_pred) < args.days:
                    logger.info(f"⚠️ Only {len(dates_in_pred)} date(s) found, fetching historical predictions...")
                    # Fetch predictions for previous days
                    for day_offset in range(args.days - 1, 0, -1):  # From oldest to newest (excluding today)
                        pred_date = end_date - pd.Timedelta(days=day_offset)
                        pred_date_str = pred_date.strftime('%Y-%m-%d')
                        
                        # Skip if we already have this date
                        if pred_date in dates_in_pred:
                            continue
                        
                        logger.info(f"🔮 Fetching predictions for {pred_date_str}...")
                        try:
                            hist_results = model.predict_with_snapshot(
                                feature_data=None,
                                snapshot_id=None if args.snapshot == "latest" else args.snapshot,
                                universe_tickers=tickers,
                                as_of_date=pred_date,
                                prediction_days=1
                            )
                            
                            if hist_results.get('success', False):
                                hist_pred = hist_results.get('predictions')
                                if hist_pred is not None and len(hist_pred) > 0:
                                    if isinstance(hist_pred, pd.Series):
                                        hist_df = hist_pred.to_frame('score')
                                    else:
                                        hist_df = hist_pred.copy()
                                    
                                    # Ensure MultiIndex
                                    if not isinstance(hist_df.index, pd.MultiIndex):
                                        hist_df.index = pd.MultiIndex.from_arrays(
                                            [[pred_date] * len(hist_df), hist_df.index],
                                            names=['date', 'ticker']
                                        )
                                    else:
                                        # Update date level
                                        new_index = pd.MultiIndex.from_arrays([
                                            [pred_date] * len(hist_df),
                                            hist_df.index.get_level_values('ticker')
                                        ], names=['date', 'ticker'])
                                        hist_df.index = new_index
                                    
                                    pred_df = pd.concat([pred_df, hist_df], axis=0)
                                    logger.info(f"✅ Added {len(hist_df)} predictions for {pred_date_str}")
                        except Exception as e:
                            logger.warning(f"Failed to fetch predictions for {pred_date_str}: {e}")
                
                all_predictions.append(pred_df)
                logger.info(f"✅ Combined predictions: {len(pred_df)} total rows")
            else:
                raise RuntimeError("No predictions returned")
        else:
            raise RuntimeError(f"Prediction failed: {results.get('error', 'Unknown error')}")
    except Exception as e:
        logger.error(f"Error generating predictions: {e}", exc_info=True)
        raise
    
    if not all_predictions:
        raise RuntimeError("No predictions generated for any date")
    
    # Combine all predictions (if we have multiple DataFrames)
    if len(all_predictions) == 1:
        combined_predictions = all_predictions[0]
    else:
        combined_predictions = pd.concat(all_predictions, axis=0)
    
    logger.info(f"✅ Combined {len(combined_predictions)} predictions")
    logger.info(f"   Date range: {combined_predictions.index.get_level_values('date').min()} to {combined_predictions.index.get_level_values('date').max()}")
    logger.info(f"   Unique tickers: {combined_predictions.index.get_level_values('ticker').nunique()}")
    
    # Apply EWMA smoothing
    logger.info("📊 Applying EWMA smoothing...")
    smoothed_predictions = calculate_ewma_smoothed_scores(
        combined_predictions,
        weights=weights_tuple,
        use_half_life=use_half_life,
        half_life_days=args.half_life if args.half_life else 3.0
    )
    
    # Generate Excel report
    logger.info("📊 Generating Excel ranking report...")
    generate_excel_ranking_report(
        smoothed_predictions,
        args.output,
        model_name="MetaRankerStacker"
    )
    
    logger.info("=" * 100)
    logger.info("✅ COMPLETE")
    logger.info("=" * 100)
    logger.info(f"Excel report: {args.output}")
    logger.info(f"Total predictions: {len(smoothed_predictions)}")
    logger.info(f"Unique tickers: {smoothed_predictions.index.get_level_values('ticker').nunique()}")
    logger.info(f"Date range: {smoothed_predictions.index.get_level_values('date').min()} to {smoothed_predictions.index.get_level_values('date').max()}")


if __name__ == "__main__":
    main()
